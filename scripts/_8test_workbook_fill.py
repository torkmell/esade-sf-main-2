#!/usr/bin/env python
"""Populate RAG_Screening_Sheet_Workbook_v1.xlsx with the 8-Test greenwashing
screening for the 20 portfolio holdings — the HOW_TO_USE.txt workflow.

Tier 1  = primary company documents in data/rag/corpus/ (the SusFin archive).
          Verbatim page-cited evidence comes from the greenwash_*.json files.
Tier 2  = external verification: data/rag/SBTi.xlsx + Latest_CP_Assessments.csv.
          SBTi / TPI status is matched by ISIN and drives dimensions 4 (Target)
          and 7 (Verification).

The workbook is the authoritative human-reviewed layer (HOW_TO_USE.txt). NB09
reads it, translates PASS/PARTIAL/FAIL/MISSING -> LOW/MED/HIGH/MISSING, and
re-generates the greenwash_*.json files from it.
"""
import glob, json, os, sys
import pandas as pd
from openpyxl import load_workbook

PROJECT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
os.chdir(PROJECT)

WB_PATH = "data/rag/RAG_Screening_Sheet_Workbook_v1.xlsx"
SHEET   = "RAG Screening Sheet"

# ── 20 holdings: ticker, company name, ISIN, corpus folder ─────────────────────
HOLDINGS = [
 ("ZURN.SW","Zurich Insurance Group Ltd","CH0011075394","09_Zurich_Insurance_Group_Ltd"),
 ("SBMO.AS","SBM Offshore NV","NL0000360618","08_SBM_Offshore_NV"),
 ("SPSN.SW","Swiss Prime Site AG","CH0010675863","13_Swiss_Prime_Site_AG"),
 ("ABBN.SW","ABB Ltd.","CH0012221716","12_ABB_Ltd"),
 ("LI.PA","Klepierre SA","FR0000121964","17_Klepierre_SA"),
 ("MRL.MC","MERLIN Properties SOCIMI, S.A.","ES0105025003","23_MERLIN_Properties_SOCIMI_SA"),
 ("GALE.SW","Galenica AG","CH0360674466","26_Galenica_AG"),
 ("EOAN.DE","E.ON SE","DE000ENAG999","24_EON_SE"),
 ("A5G.IR","AIB Group plc","IE00BF0L3536","06_AIB_Group_plc"),
 ("UCB.BR","UCB S.A.","BE0003739530","14_UCB_SA"),
 ("AZN.L","AstraZeneca PLC","GB0009895292","22_AstraZeneca_PLC"),
 ("AGN.AS","Aegon Ltd.","NL0000303709","05_Aegon_Ltd"),
 ("LLOY.L","Lloyds Banking Group plc","GB0008706128","07_Lloyds_Banking_Group_plc"),
 ("ALFA.ST","Alfa Laval AB","SE0000695876","02_Alfa_Laval_AB"),
 ("NHY.OL","Norsk Hydro ASA","NO0005052605","03_Norsk_Hydro_ASA"),
 ("TEL2-B.ST","Tele2 AB Class B","SE0005190238","33_Tele2_AB_Class_B"),
 ("ITX.MC","Industria de Diseno Textil, S.A.","ES0148396007","27_Industria_de_Diseno_Textil_SA"),
 ("ORNBV.HE","Orion Oyj Class B","FI0009014377","28_Orion_Oyj_Class_B"),
 ("SOBI.ST","Swedish Orphan Biovitrum AB","SE0000872095","32_Swedish_Orphan_Biovitrum_AB"),
 ("SUBC.OL","Subsea 7 S.A.","LU0075646355","40_Subsea_7_SA"),
]
# Bloomberg-style ticker for the workbook Ticker column (new rows only)
BBG_TICKER = {"LI.PA":"LI","MRL.MC":"MRL","GALE.SW":"GALE","A5G.IR":"A5G",
              "AGN.AS":"AGN","LLOY.L":"LLOY","ITX.MC":"ITX"}
# Evidence-confidence override where the source is thin (1-5 scale)
EV_CONF = {"SPSN.SW":3, "SUBC.OL":3}

DIMS = ["specificity","metric","baseline","target","time_horizon",
        "scope","verification","consistency"]
TRANSLATE = {"LOW":"PASS","MED":"PARTIAL","HIGH":"FAIL","MISSING":"MISSING"}

# ── Tier-2 data (built by the earlier extraction step) ─────────────────────────
t2 = pd.read_csv("outputs/rag/_tier2_sbti_cp.csv").set_index("ticker")

def tier2_revise(dim, t1_rating, sbti_status, tpi_aligned):
    """Apply SBTi / TPI evidence to dimensions 4 (Target) and 7 (Verification).
    Other dimensions keep their Tier-1 rating. Scale: LOW/MED/HIGH/MISSING."""
    if dim == "target":
        if sbti_status == "Targets set":          return "LOW"   # SBTi-validated
        if tpi_aligned:                            return "LOW"   # TPI 1.5C-aligned
        if sbti_status in ("Committed", "Commitment removed"):
            return "MED"                                          # not (yet) validated
        return "MED" if t1_rating == "LOW" else t1_rating         # no external check
    if dim == "verification":
        if sbti_status == "Targets set":          return "LOW"   # registry-confirmed
        if tpi_aligned:                            return "LOW"   # TPI-assessed
        return "MED" if t1_rating in ("LOW","MED") else t1_rating
    return t1_rating

def primary_doc(folder):
    mf = f"data/rag/corpus/{folder}/_8test/manifest.txt"
    line = open(mf, encoding="utf-8").readline().strip()
    return line.split(":",1)[1].strip() if ":" in line else line

# ── Build the per-company assessment ───────────────────────────────────────────
records = []
for tk, name, isin, folder in HOLDINGS:
    jpath = f"outputs/rag/greenwash_{tk}.json"
    if not os.path.exists(jpath):
        sys.exit(f"Missing Tier-1 JSON: {jpath}")
    j = json.load(open(jpath))
    row = t2.loc[tk]
    sbti_status = str(row["sbti_near_term_status"])
    sbti_year   = str(row["sbti_near_term_year"])
    sbti_class  = str(row["sbti_near_term_class"])
    tpi_raw     = str(row["tpi_cp"])
    tpi_aligned = "degree" in tpi_raw.lower() or "1.5" in tpi_raw
    in_sbti     = sbti_status not in ("--", "nan", "")

    ratings = {}
    for d in DIMS:
        t1 = j["dimensions"][d]["rating"].strip().upper()
        ratings[d] = TRANSLATE[tier2_revise(d, t1, sbti_status, tpi_aligned)]

    n_fail    = sum(v == "FAIL" for v in ratings.values())
    n_partial = sum(v == "PARTIAL" for v in ratings.values())
    n_missing = sum(v == "MISSING" for v in ratings.values())
    verdict = ("EXCLUDE" if n_fail >= 3 else
               "WATCHLIST" if n_fail == 2 else "PASS")

    sp  = j["dimensions"]["specificity"]
    doc = primary_doc(folder)

    # SBTi cell text + Tier-2 sentence for the rationale
    if not in_sbti:
        sbti_cell = "Not in SBTi database (checked 2026-05-21)"
        t2_sent = "SBTi: no validated target in the SBTi database."
    else:
        sbti_cell = sbti_status + (f" ({sbti_class})" if sbti_class not in ("nan","") else "")
        t2_sent = f"SBTi: {sbti_status}" + (
            f", {sbti_class} near-term target ({sbti_year})." if sbti_status == "Targets set"
            else ".")
    if tpi_aligned:
        t2_sent += f" TPI Carbon Performance: {tpi_raw}-aligned."

    note = j.get("analyst_note","").strip()
    if len(note) > 240:
        note = note[:237].rsplit(" ",1)[0] + "..."
    rationale = (f"Main claim: \"{sp['quote'][:140]}\" ({doc}, p{sp['page']}). "
                 f"{t2_sent} 8-Test: {n_partial} PARTIAL, {n_fail} FAIL, "
                 f"{n_missing} MISSING -> {verdict}. {note}")

    records.append(dict(
        ticker=tk, name=name, isin=isin,
        sbti_checked="Yes", tpi_checked="Yes",
        sbti_status=sbti_cell,
        sbti_year=(sbti_year if sbti_status == "Targets set" and sbti_year not in ("nan","") else ""),
        main_claim=sp["quote"], claim_doc=doc, claim_page=sp["page"],
        r=ratings,
        key_quote=sp["quote"], key_doc=doc, key_page=sp["page"],
        ev_conf=EV_CONF.get(tk,4), ai_conf=4,
        verif_status="PENDING",
        rationale=rationale, verdict=verdict,
        n_fail=n_fail, n_partial=n_partial, n_missing=n_missing))

# ── Write into the workbook ────────────────────────────────────────────────────
# Columns A/B/C (Company Name, Ticker, ISIN) are formulas pulling from the
# Universe Tracker, so they are matched on their CACHED values (data_only=True);
# the assessment columns are written to the formula-preserving workbook.
wbv = load_workbook(WB_PATH, data_only=True)
wsv = wbv[SHEET]
wb  = load_workbook(WB_PATH)
ws  = wb[SHEET]

HEADER_ROW = 6
hdr = {}
for c in range(1, ws.max_column + 1):
    v = ws.cell(HEADER_ROW, c).value
    if v is not None:
        hdr[str(v).strip()] = c

def col(label_prefix):
    for k, c in hdr.items():
        if k.startswith(label_prefix):
            return c
    sys.exit(f"Workbook column not found: {label_prefix!r}")

C_NAME = col("Company Name"); C_TKR = col("Ticker"); C_ISIN = col("ISIN")
C_SBTICHK = col("SBTi Database Checked"); C_TPICHK = col("TPI Checked")
C_SBTIST = col("SBTi Status"); C_SBTIYR = col("SBTi Approval Year")
C_CLAIM = col("Main Claim Tested"); C_CLAIMDOC = col("Claim Source Document")
C_CLAIMPG = col("Claim Page")
C_T = [col(f"8-Test {i}:") for i in range(1, 9)]
C_KQ = col("Key Evidence Quote"); C_KDOC = col("Key Evidence Source Doc")
C_KPG = col("Key Evidence Page Ref")
C_EVCONF = col("Evidence Confidence"); C_AICONF = col("AI Extraction Confidence")
C_VST = col("Verification Status"); C_RAT = col("Verdict Rationale")

# Map existing rows by ISIN — read CACHED values (columns A/B/C are formulas)
last = HEADER_ROW
isin_row = {}
for r in range(HEADER_ROW + 1, wsv.max_row + 1):
    nm = wsv.cell(r, C_NAME).value
    iv = wsv.cell(r, C_ISIN).value
    if nm not in (None, ""):
        last = r
    if iv not in (None, ""):
        isin_row[str(iv).strip().upper()] = r

def put(r, rec):
    # ISIN is written as a literal value (the workbook's column C is a formula
    # pulling from Universe Tracker; openpyxl drops cached formula values on
    # save, so a literal keeps the row machine-readable for NB09). The literal
    # equals the formula's output — the data is unchanged. Columns A/B
    # (Company Name / Ticker) are left as-is per HOW_TO_USE.txt.
    ws.cell(r, C_ISIN, rec["isin"])
    ws.cell(r, C_SBTICHK, rec["sbti_checked"])
    ws.cell(r, C_TPICHK,  rec["tpi_checked"])
    ws.cell(r, C_SBTIST,  rec["sbti_status"])
    ws.cell(r, C_SBTIYR,  rec["sbti_year"])
    ws.cell(r, C_CLAIM,   rec["main_claim"])
    ws.cell(r, C_CLAIMDOC,rec["claim_doc"])
    ws.cell(r, C_CLAIMPG, rec["claim_page"])
    for i, d in enumerate(DIMS):
        ws.cell(r, C_T[i], rec["r"][d])
    ws.cell(r, C_KQ,   rec["key_quote"])
    ws.cell(r, C_KDOC, rec["key_doc"])
    ws.cell(r, C_KPG,  rec["key_page"])
    ws.cell(r, C_EVCONF, rec["ev_conf"])
    ws.cell(r, C_AICONF, rec["ai_conf"])
    ws.cell(r, C_VST,  rec["verif_status"])
    ws.cell(r, C_RAT,  rec["rationale"])

new_row = last
existing, added = 0, 0
print(f"{'TICKER':<10}{'ROW':<7}{'PARTIAL':<9}{'FAIL':<6}{'MISSING':<9}VERDICT")
print("-" * 56)
for rec in records:
    r = isin_row.get(rec["isin"].upper())
    if r:
        existing += 1
        tag = "exist"
    else:
        new_row += 1
        r = new_row
        ws.cell(r, C_NAME, rec["name"])
        ws.cell(r, C_TKR,  BBG_TICKER.get(rec["ticker"], rec["ticker"]))
        ws.cell(r, C_ISIN, rec["isin"])
        added += 1
        tag = "NEW"
    put(r, rec)
    print(f"{rec['ticker']:<10}{r:<3}{tag:<4}{rec['n_partial']:<9}"
          f"{rec['n_fail']:<6}{rec['n_missing']:<9}{rec['verdict']}")

wb.save(WB_PATH)
print("-" * 56)
print(f"Workbook updated: {WB_PATH}")
print(f"  existing rows filled: {existing}   new rows appended: {added}")
print(f"  total holdings screened: {len(records)}")
